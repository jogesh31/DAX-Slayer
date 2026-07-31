"""
Local web backend for DAX Slayer — Power BI dependency analysis & cleanup.

Run via `Run Dependency Analyzer.bat` (or `python app.py`), then open
http://127.0.0.1:8765 — or launch it straight from Power BI Desktop's
External Tools ribbon once pbitool.json is installed (see SETUP.md).
"""
from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
import webbrowser
import zipfile
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory

import pbi_connect
from dax_explain import explain_expression, find_function_calls
from dax_format import format_expression
from dax_functions import all_functions
from dax_snippets import all_snippets
from graph import build_graph, dependents_of

app = Flask(__name__, static_folder="static", static_url_path="")

STATE = {
    "conn": None,
    "port": None,
    "report_folder": None,
    "temp_pbix_extract": None,  # temp folder created from .pbix extraction, cleaned up on shutdown
    "tables_df": None,
    "columns_df": None,
    "measures_df": None,
    "relationships_df": None,
    "graph": None,
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)  # "TL dashboard" folder
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)


def _extract_pbix_report_folder(pbix_path: str) -> str | None:
    """A .pbix is a ZIP archive with the same Report/ structure inside as a .pbip
    project has on disk. Extract it to a temp folder and return the path to the
    extracted Report/ folder, or None if extraction fails.

    The extracted folder is tracked in STATE so it can be cleaned up on shutdown."""
    if not pbix_path or not pbix_path.lower().endswith(".pbix"):
        return None
    if not os.path.isfile(pbix_path):
        return None
    try:
        if STATE.get("temp_pbix_extract") and os.path.isdir(STATE["temp_pbix_extract"]):
            shutil.rmtree(STATE["temp_pbix_extract"], ignore_errors=True)

        temp_root = tempfile.mkdtemp(prefix="dax_slayer_pbix_")
        with zipfile.ZipFile(pbix_path, "r") as zf:
            zf.extractall(temp_root)

        report_folder = os.path.join(temp_root, "Report")
        if os.path.isdir(report_folder):
            STATE["temp_pbix_extract"] = temp_root
            return report_folder
        shutil.rmtree(temp_root, ignore_errors=True)
        return None
    except (zipfile.BadZipFile, OSError):
        return None


def _err(e: Exception, code: int = 500):
    traceback.print_exc()
    return jsonify({"error": str(e)}), code


def _search_roots() -> list[str]:
    roots = [PROJECT_ROOT]
    home = os.path.expanduser("~")
    for sub in ("Documents", "Desktop", os.path.join("OneDrive", "Documents"), os.path.join("OneDrive", "Desktop")):
        p = os.path.join(home, sub)
        if os.path.isdir(p) and p not in roots:
            roots.append(p)
    return roots


def _find_report_folders(max_depth: int = 2) -> list[str]:
    """Scan a handful of likely project locations for PBIP `*.Report` folders,
    depth-limited so this stays fast even under a large Documents folder."""
    found: list[str] = []
    for root in _search_roots():
        root_depth = root.rstrip(os.sep).count(os.sep)
        for dirpath, dirnames, _files in os.walk(root):
            depth = dirpath.rstrip(os.sep).count(os.sep) - root_depth
            if depth >= max_depth:
                dirnames[:] = []
                continue
            for d in list(dirnames):
                if d.endswith(".Report"):
                    found.append(os.path.join(dirpath, d))
        if len(found) > 200:  # safety valve on huge/unusual trees
            break
    return sorted(set(found))


def _clean_title(title: str | None) -> str:
    if not title:
        return ""
    return re.sub(r"\s*-\s*Power BI Desktop\s*$", "", title).strip()


def _best_report_folder_match(report_title: str, folders: list[str]) -> str | None:
    clean = _clean_title(report_title)
    if not clean or not folders:
        return None
    names = {os.path.splitext(os.path.basename(f))[0]: f for f in folders}
    matches = difflib.get_close_matches(clean, names.keys(), n=1, cutoff=0.4)
    return names[matches[0]] if matches else None


@app.get("/")
def index():
    return send_from_directory("static", "index.html")


@app.get("/api/state")
def api_state():
    return jsonify({
        "connected": STATE["conn"] is not None,
        "port": STATE["port"],
        "report_folder": STATE.get("report_folder"),
    })


@app.get("/api/instances")
def api_instances():
    try:
        instances = pbi_connect.discover_local_instances()
        report_folders = _find_report_folders()
        return jsonify({
            "instances": [
                {
                    "port": i.port,
                    "process_id": i.process_id,
                    "report_title": _clean_title(i.report_title) or f"Untitled report (port {i.port})",
                    # a .pbip project's own sibling .Report folder, read straight off
                    # Power BI Desktop's command line — auto-detected, zero guessing
                    "auto_report_folder": i.report_folder,
                    "file_path": i.file_path,
                    "is_pbix": i.is_pbix,
                    # fallback for when the file path couldn't be read at all
                    "suggested_report_folder": i.report_folder or _best_report_folder_match(i.report_title, report_folders),
                }
                for i in instances
            ],
            "report_folders": [os.path.basename(p) for p in report_folders],
            "report_folder_full_paths": report_folders,
        })
    except Exception as e:
        return _err(e)


@app.post("/api/connect")
def api_connect():
    try:
        body = request.get_json(force=True) or {}
        port = body.get("port")
        report_folder = body.get("report_folder")

        if STATE["conn"] is not None:
            try:
                STATE["conn"].close()
            except Exception:
                pass

        if port:
            conn = pbi_connect.PowerBIConnection(int(port))
        else:
            conn = pbi_connect.connect_first_available()

        STATE["conn"] = conn
        STATE["port"] = conn.port
        STATE["report_folder"] = report_folder

        # If no report_folder was provided, try to auto-detect:
        # Look up the instance and see if it's a .pbix we can extract, or a .pbip with a Report folder
        if not report_folder:
            instances = pbi_connect.discover_local_instances()
            matching = next((i for i in instances if i.port == conn.port), None)
            if matching:
                if matching.report_folder:
                    STATE["report_folder"] = matching.report_folder
                elif matching.is_pbix and matching.file_path:
                    extracted = _extract_pbix_report_folder(matching.file_path)
                    if extracted:
                        STATE["report_folder"] = extracted

        return jsonify({"connected": True, "port": conn.port, "report_folder": STATE.get("report_folder")})
    except Exception as e:
        return _err(e)


def _require_conn() -> pbi_connect.PowerBIConnection:
    if STATE["conn"] is None:
        raise RuntimeError("Not connected. Call /api/connect first.")
    return STATE["conn"]


@app.post("/api/analyze")
def api_analyze():
    try:
        conn = _require_conn()
        tables_df = conn.list_tables()
        columns_df = conn.list_columns()
        measures_df = conn.list_measures()
        relationships_df = conn.list_relationships()

        STATE["tables_df"] = tables_df
        STATE["columns_df"] = columns_df
        STATE["measures_df"] = measures_df
        STATE["relationships_df"] = relationships_df

        report_folder = STATE.get("report_folder")
        mgraph = build_graph(tables_df, columns_df, measures_df, report_folder)
        STATE["graph"] = mgraph

        nodes_out = []
        for nid, n in mgraph.nodes.items():
            entry = {
                "id": nid,
                "type": n.node_type,
                "table": n.table,
                "name": n.name,
                "isHidden": n.is_hidden,
            }
            if n.node_type == "column":
                entry["isCalculated"] = n.is_calculated
            if n.node_type == "measure":
                entry["displayFolder"] = n.display_folder
            if n.node_type in ("measure", "column"):
                u = mgraph.usage[nid]
                entry["usageLevel"] = u.level
                entry["usedByMeasures"] = u.used_by_measures
                entry["usedInReport"] = u.used_in_report
                entry["reportSources"] = u.report_sources
            if n.node_type in ("measure", "column"):
                entry["expression"] = n.expression
            nodes_out.append(entry)

        edges_out = [{"source": e.source, "target": e.target} for e in mgraph.edges]

        rel_out = []
        for _, r in relationships_df.iterrows():
            rel_out.append({
                "fromTable": r["FromTable"], "fromColumn": r["FromColumn"],
                "toTable": r["ToTable"], "toColumn": r["ToColumn"],
                "isActive": bool(r["IsActive"]),
            })

        unused = [n for n in nodes_out if n.get("usageLevel") == "unused"]

        return jsonify({
            "nodes": nodes_out,
            "edges": edges_out,
            "relationships": rel_out,
            "unresolvedRefs": mgraph.unresolved_refs,
            "circular": mgraph.circular,
            "reportFilesScanned": mgraph.report_files_scanned,
            "summary": {
                "tableCount": len(tables_df),
                "columnCount": len(columns_df),
                "measureCount": len(measures_df),
                "relationshipCount": len(relationships_df),
                "unusedCount": len(unused),
            },
        })
    except Exception as e:
        return _err(e)


@app.get("/api/dependents")
def api_dependents():
    try:
        node_id = request.args.get("id")
        mgraph = STATE.get("graph")
        if mgraph is None:
            raise RuntimeError("Run /api/analyze first.")
        return jsonify({"dependents": dependents_of(mgraph, node_id)})
    except Exception as e:
        return _err(e)


@app.get("/api/dax-functions")
def api_dax_functions():
    """The full offline DAX function reference library — fetched once by the
    frontend and cached client-side, since it's static data."""
    try:
        return jsonify({
            "functions": [
                {"name": f.name, "category": f.category, "description": f.description, "syntax": f.syntax, "example": f.example}
                for f in all_functions()
            ]
        })
    except Exception as e:
        return _err(e)


@app.get("/api/dax-snippets")
def api_dax_snippets():
    """"Most Used DAX" — a curated library of common patterns (Date table,
    time intelligence, ranking, etc.), not tied to whatever model happens to
    be connected. functionsUsed is precomputed per snippet so the frontend
    can reuse the same function-name highlighting/popover as everywhere
    else in the tool."""
    try:
        out = []
        for s in all_snippets():
            uses = find_function_calls(s.dax)
            expr_uses = find_function_calls(s.expression)
            out.append({
                "name": s.name,
                "category": s.category,
                "description": s.description,
                "dax": s.dax,
                "functionsUsed": [{"name": u.name, "start": u.start, "end": u.end} for u in uses],
                "kind": s.kind,
                "defaultObjectName": s.default_object_name,
                "expression": s.expression,
                "expressionFunctionsUsed": [{"name": u.name, "start": u.start, "end": u.end} for u in expr_uses],
            })
        return jsonify({"snippets": out})
    except Exception as e:
        return _err(e)


@app.post("/api/deploy-snippet")
def api_deploy_snippet():
    """Creates a brand-new measure, calculated column, or calculated table
    from an edited "Most Used DAX" template — the user has already renamed
    it and swapped in their own table/column references. Same backup-first
    safety as every other write path here."""
    try:
        conn = _require_conn()
        body = request.get_json(force=True) or {}
        kind = body.get("kind")
        table = (body.get("table") or "").strip()
        name = (body.get("name") or "").strip()
        expression = (body.get("expression") or "").strip()

        if kind not in ("measure", "column", "table"):
            raise RuntimeError(f"Unknown snippet kind: '{kind}'.")
        if kind != "table" and not table:
            raise RuntimeError("Pick a target table.")
        if not name:
            raise RuntimeError("Give it a name.")
        if not expression:
            raise RuntimeError("The expression is empty.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"snapshot_{timestamp}.bim")
        conn.export_bim_snapshot(backup_path)

        if kind == "measure":
            conn.create_measure(table, name, expression)
        elif kind == "column":
            conn.create_column(table, name, expression)
        else:
            conn.create_table(name, expression)

        return jsonify({"deployed": True, "backup": backup_path, "kind": kind, "table": table, "name": name})
    except Exception as e:
        return _err(e)


def _build_resolver(mgraph):
    """A `(table, name) -> friendly phrase` function grounded in the actual
    semantic model, so the explainer can say "the Total Teachers Completed
    measure" or "the COURSE_ID column in VW_TPD_AVG_MODULE_COMPLETION"
    instead of echoing raw `Table[Column]` bracket syntax back at the
    reader. This is the "use the whole semantic model" part — the graph
    already knows every table/column/measure, so resolution is a lookup,
    not a guess."""
    def resolve(table, name):
        if table:
            node = mgraph.nodes.get(f"{table}[{name}]")
            if node and node.node_type == "measure":
                return f"the {name} measure"
            if node and node.node_type == "column":
                return f"the {name} column in {table}"
            return f"{table}[{name}]"
        # bare [Name] — ambiguous; DAX itself resolves measures first
        for node in mgraph.nodes.values():
            if node.node_type == "measure" and node.name == name:
                return f"the {name} measure"
        for node in mgraph.nodes.values():
            if node.node_type == "column" and node.name == name:
                return f"the {name} column"
        return f"[{name}]"
    return resolve


@app.get("/api/explain")
def api_explain():
    """Rule-based, offline plain-English breakdown of one measure's DAX —
    no network call, no AI, just a structural walk of the expression
    grounded against the real semantic model."""
    try:
        node_id = request.args.get("id")
        mgraph = STATE.get("graph")
        if mgraph is None:
            raise RuntimeError("Run /api/analyze first.")
        node = mgraph.nodes.get(node_id)
        if node is None:
            raise RuntimeError(f"Unknown object: {node_id}")
        return jsonify(explain_expression(node.expression or "", resolve=_build_resolver(mgraph)))
    except Exception as e:
        return _err(e)


@app.get("/api/format")
def api_format():
    """Beautified (indented, multi-line) version of a measure's DAX for
    display — purely cosmetic, doesn't touch the live model. See
    /api/format-deploy to actually write the reformatted text back."""
    try:
        node_id = request.args.get("id")
        mgraph = STATE.get("graph")
        if mgraph is None:
            raise RuntimeError("Run /api/analyze first.")
        node = mgraph.nodes.get(node_id)
        if node is None:
            raise RuntimeError(f"Unknown object: {node_id}")
        formatted = format_expression(node.expression or "")
        # Positions must be recomputed against the formatted text (reflowing
        # onto multiple lines shifts every offset) so the box can highlight
        # function names correctly in what's actually displayed.
        uses = find_function_calls(formatted)
        return jsonify({
            "formatted": formatted,
            "functionsUsed": [{"name": u.name, "start": u.start, "end": u.end} for u in uses],
        })
    except Exception as e:
        return _err(e)


@app.post("/api/format-deploy")
def api_format_deploy():
    """Reformats a measure's DAX and writes the beautified text back to the
    live model — same backup-first safety as every other write path here."""
    try:
        conn = _require_conn()
        body = request.get_json(force=True) or {}
        node_id = body.get("id")
        mgraph = STATE.get("graph")
        if mgraph is None:
            raise RuntimeError("Run /api/analyze first.")
        node = mgraph.nodes.get(node_id)
        if node is None or node.node_type != "measure":
            raise RuntimeError(f"Unknown measure: {node_id}")

        formatted = format_expression(node.expression or "")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"snapshot_{timestamp}.bim")
        conn.export_bim_snapshot(backup_path)

        conn.set_measure_expression(node.table, node.name, formatted)

        return jsonify({"deployed": True, "backup": backup_path, "formatted": formatted})
    except Exception as e:
        return _err(e)


@app.post("/api/deploy")
def api_deploy():
    """Apply a batch of pending deletions to the live model.

    Always writes a full BIM snapshot to backups/ before touching anything,
    then applies all deletions as one TOM transaction — if SaveChanges
    throws partway through, the live model is left untouched (TOM only
    commits on a clean SaveChanges), and the snapshot is on disk regardless
    so a manual restore is possible even if something else goes wrong.
    """
    try:
        conn = _require_conn()
        body = request.get_json(force=True) or {}
        deletions = body.get("deletions", [])
        if not deletions:
            return jsonify({"error": "No deletions provided."}), 400

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"snapshot_{timestamp}.bim")
        conn.export_bim_snapshot(backup_path)

        conn.delete_objects(deletions)

        return jsonify({"deployed": True, "backup": backup_path, "count": len(deletions)})
    except Exception as e:
        return _err(e)


@app.post("/api/organize-measures")
def api_organize_measures():
    """Move every measure in the model into a single DisplayFolder, so the
    Power BI field list stops listing them loose at the table root. Same
    backup-first safety as /api/deploy — a snapshot is written before the
    live model is touched."""
    try:
        conn = _require_conn()
        body = request.get_json(force=True) or {}
        folder_name = (body.get("folder_name") or "Measures").strip()
        if not folder_name:
            return jsonify({"error": "Folder name can't be empty."}), 400
        scope = body.get("measures")  # None = every measure in the model

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"snapshot_{timestamp}.bim")
        conn.export_bim_snapshot(backup_path)

        count = conn.set_measures_folder(folder_name, scope)

        return jsonify({"organized": True, "backup": backup_path, "count": count, "folderName": folder_name})
    except Exception as e:
        return _err(e)


@app.get("/api/export-cleanup-report")
def api_export_cleanup_report():
    """A downloadable, human-readable cleanup report — separate sheets for
    used/unused measures and used/unused calculated columns, since that's
    what someone doing cleanup actually needs to hand to a reviewer or
    work through offline. Replaces the old silent server-side JSON/CSV dump
    (which wrote a file nobody could easily get to) with a real download."""
    try:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from flask import send_file

        mgraph = STATE.get("graph")
        if mgraph is None:
            raise RuntimeError("Run /api/analyze first.")

        def usage_detail(nid: str) -> str:
            u = mgraph.usage.get(nid)
            if u is None or u.level == "unused":
                return "Not used anywhere"
            if u.level == "used-in-report":
                return "Used on a report visual" + (f" ({', '.join(u.report_sources[:3])})" if u.report_sources else "")
            return "Used by other measures only"

        def used_by_names(nid: str) -> str:
            u = mgraph.usage.get(nid)
            if not u or not u.used_by_measures:
                return ""
            names = [mgraph.nodes[i].name for i in u.used_by_measures if i in mgraph.nodes]
            return ", ".join(names)

        measures = [n for n in mgraph.nodes.values() if n.node_type == "measure"]
        calc_cols = [n for n in mgraph.nodes.values() if n.node_type == "column" and n.is_calculated]

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4D7FFF")

        def add_sheet(title: str, headers: list[str], rows: list[list]):
            ws = wb.create_sheet(title)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")
            for row in rows:
                ws.append(row)
            ws.freeze_panes = "A2"
            for i, h in enumerate(headers, start=1):
                width = max(len(h), *(len(str(r[i - 1])) for r in rows)) if rows else len(h)
                ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 60)
            return ws

        used_measures = sorted((n for n in measures if mgraph.usage[n.id].level != "unused"), key=lambda n: (n.table, n.name))
        add_sheet(
            "Used Measures",
            ["Table", "Name", "Display Folder", "Usage Detail"],
            [[n.table, n.name, n.display_folder or "", usage_detail(n.id)] for n in used_measures],
        )

        unused_measures = sorted((n for n in measures if mgraph.usage[n.id].level == "unused"), key=lambda n: (n.table, n.name))
        add_sheet(
            "Unused Measures",
            ["Table", "Name", "Display Folder", "Status"],
            [[n.table, n.name, n.display_folder or "", "Safe to delete — not used anywhere"] for n in unused_measures],
        )

        used_cols = sorted((n for n in calc_cols if mgraph.usage[n.id].level != "unused"), key=lambda n: (n.table, n.name))
        add_sheet(
            "Used Calculated Columns",
            ["Table", "Name", "Usage Detail", "Used By Measures"],
            [[n.table, n.name, usage_detail(n.id), used_by_names(n.id)] for n in used_cols],
        )

        unused_cols = sorted((n for n in calc_cols if mgraph.usage[n.id].level == "unused"), key=lambda n: (n.table, n.name))
        add_sheet(
            "Unused Calculated Columns",
            ["Table", "Name", "Status"],
            [[n.table, n.name, "Safe to delete — not used anywhere"] for n in unused_cols],
        )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"dax_slayer_cleanup_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return _err(e)


def _auto_connect_from_args(server: str | None, database: str | None):
    """When launched from Power BI Desktop's External Tools ribbon, %server%
    is substituted with something like "localhost:56789" — pull the port out
    and connect immediately so the app opens already attached to the report
    that invoked it, instead of showing an empty instance picker."""
    if not server:
        return
    m = re.search(r":(\d+)", server)
    if not m:
        return
    port = int(m.group(1))
    try:
        conn = pbi_connect.PowerBIConnection(port)
        STATE["conn"] = conn
        STATE["port"] = conn.port
        instances = pbi_connect.discover_local_instances()
        this_instance = next((i for i in instances if i.port == port), None)
        if this_instance:
            if this_instance.report_folder:
                STATE["report_folder"] = this_instance.report_folder
            elif this_instance.is_pbix and this_instance.file_path:
                extracted = _extract_pbix_report_folder(this_instance.file_path)
                if extracted:
                    STATE["report_folder"] = extracted
            elif not this_instance.is_pbix:
                report_folders = _find_report_folders()
                STATE["report_folder"] = _best_report_folder_match(this_instance.report_title, report_folders)
        print(f"Auto-connected to Power BI Desktop on port {port}, report_folder={STATE['report_folder']}")
    except Exception as e:
        print(f"Auto-connect failed ({e}); use the picker in the UI instead.")


def _cleanup():
    if STATE.get("temp_pbix_extract") and os.path.isdir(STATE["temp_pbix_extract"]):
        shutil.rmtree(STATE["temp_pbix_extract"], ignore_errors=True)
        print(f"Cleaned up temp extraction: {STATE['temp_pbix_extract']}")


if __name__ == "__main__":
    import atexit
    atexit.register(_cleanup)

    parser = argparse.ArgumentParser()
    parser.add_argument("--server", default=None, help="XMLA server string passed by Power BI's External Tools (e.g. localhost:56789)")
    parser.add_argument("--database", default=None, help="Database name passed by Power BI's External Tools")
    parser.add_argument("--no-browser", action="store_true")
    args = parser.parse_args()

    _auto_connect_from_args(args.server, args.database)

    web_port = 8765
    url = f"http://127.0.0.1:{web_port}"
    if not args.no_browser:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    app.run(host="127.0.0.1", port=web_port, debug=False)
